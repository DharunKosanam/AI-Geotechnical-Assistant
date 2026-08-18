"""
Thread-wide highlight export (HIGHLIGHTS_ENABLED): pure builders, no I/O.

Rows are ordered by message (thread order), then by position within the
message. Each row carries a CONTEXT snippet: the surrounding SOURCE text of
the message (messages.content, UTF-16 offsets), roughly +/-150 code units on
each side, trimmed to a sentence boundary when one falls inside the window,
else to a word boundary, ellipsised where text was omitted, lightly stripped
of markdown block/inline markers and whitespace-collapsed so it reads as one
line of prose. The highlighted span itself is shown inside the context as the
rendered selectedText: bold in Markdown, [bracketed] in Excel.

Markdown output escapes every user/model-supplied string it prints as text
(highlighted text, notes, context, title) so it renders literally; multi-line
notes keep their lines via hard breaks (two trailing spaces).

Excel output reuses the GeoPilot generic workbook builder
(app.workspace.export.build_workbook: one table + summary sheet) unchanged,
then a post-pass here sets wrap-text and a readable width on the long text
columns. The GeoPilot exporter itself is not modified.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.workspace.export import build_workbook

CONTEXT_RADIUS = 150  # UTF-16 code units on each side of the span
ELLIPSIS = "…"

XLSX_COLUMNS = ["Date", "Colour", "Highlighted text", "Note", "Context", "Thread title"]
XLSX_WRAP_COLUMNS = (3, 4, 5)  # Highlighted text, Note, Context (1-based)
XLSX_WRAP_WIDTH = 60
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MD_MEDIA_TYPE = "text/markdown; charset=utf-8"


# ---------------------------------------------------------------------------
# UTF-16 helpers (JS string offsets; same semantics as routers/highlights.py)
# ---------------------------------------------------------------------------

def _u16_len(text: str) -> int:
    return len(text.encode("utf-16-le")) // 2


def _u16_slice(text: str, start: int, end: int) -> str:
    start = max(0, start)
    end = max(start, end)
    return text.encode("utf-16-le")[2 * start:2 * end].decode("utf-16-le", errors="replace")


# ---------------------------------------------------------------------------
# Rows
# ---------------------------------------------------------------------------

@dataclass
class ExportRow:
    highlight_id: str
    message_id: str
    message_index: int          # position of the message in the thread (large if unknown)
    start_offset: int
    end_offset: int
    selected_text: str
    note: str
    colour: str
    created_at: Optional[datetime]
    context_before: str         # cleaned prose before the span ("" if none)
    context_after: str          # cleaned prose after the span
    context_available: bool     # False when the message row could not be read


def build_rows(
    highlights: Iterable[Dict[str, Any]],
    messages: Sequence[Dict[str, Any]],
) -> List[ExportRow]:
    """Order + contextualise. ``messages`` are the thread's message docs in
    thread order (each with ``_id`` and ``content``); highlights whose message
    is missing sort last and get no context."""
    by_id: Dict[str, Tuple[int, Any]] = {}
    for idx, m in enumerate(messages):
        by_id[str(m.get("_id"))] = (idx, m.get("content"))
    rows: List[ExportRow] = []
    for h in highlights:
        idx, content = by_id.get(str(h.get("messageId")), (10**9, None))
        start = int(h.get("startOffset") or 0)
        end = int(h.get("endOffset") or 0)
        if isinstance(content, str):
            before, after = context_window(content, start, end)
            available = True
        else:
            before, after, available = "", "", False
        rows.append(
            ExportRow(
                highlight_id=str(h.get("_id") or h.get("id") or ""),
                message_id=str(h.get("messageId") or ""),
                message_index=idx,
                start_offset=start,
                end_offset=end,
                selected_text=str(h.get("selectedText") or ""),
                note=str(h.get("note") or ""),
                colour=str(h.get("colour") or ""),
                created_at=h.get("createdAt") if isinstance(h.get("createdAt"), datetime) else None,
                context_before=before,
                context_after=after,
                context_available=available,
            )
        )
    rows.sort(key=lambda r: (r.message_index, r.start_offset, r.created_at or datetime.min))
    return rows


# ---------------------------------------------------------------------------
# Context window
# ---------------------------------------------------------------------------

# A sentence end is . ! ? followed by whitespace (not the "1." of a list marker),
# or a paragraph break.
_SENTENCE_END_BEFORE = re.compile(r"(?:(?<![\d\s])[.!?]\s+|\n\s*\n)")
_SENTENCE_END_AFTER = re.compile(r"(?<![\d\s])[.!?](?=\s|$)|\n\s*\n")


def context_window(content: str, start: int, end: int, radius: int = CONTEXT_RADIUS) -> Tuple[str, str]:
    """Return (before, after) prose around content[start:end] (UTF-16 offsets)."""
    total = _u16_len(content)
    lo = max(0, start - radius)
    hi = min(total, end + radius)
    before = _u16_slice(content, lo, start)
    after = _u16_slice(content, end, hi)

    omitted_before = lo > 0
    omitted_after = hi < total
    if omitted_before:
        # Prefer to start at a sentence/paragraph boundary inside the window,
        # else at a word boundary; never mid-word.
        # Latest sentence/paragraph boundary that still leaves some prose
        # (a boundary right before a list marker would leave only "1. **").
        ends = [m.end() for m in _SENTENCE_END_BEFORE.finditer(before) if m.end() < len(before)]
        chosen = next((e for e in reversed(ends) if _clean_context(before[e:])), None)
        if chosen is not None:
            before = before[chosen:]
        else:
            ws = before.find(" ")
            nl = before.find("\n")
            cut = min(x for x in (ws, nl) if x >= 0) if (ws >= 0 or nl >= 0) else -1
            if 0 <= cut < len(before) - 1:
                before = before[cut + 1:]
    if omitted_after:
        m = _SENTENCE_END_AFTER.search(after)
        if m is not None and m.start() > 0:
            after = after[: m.end() if m.group(0)[0] in ".!?" else m.start()]
        else:
            cut = max(after.rfind(" "), after.rfind("\n"))
            if cut > 0:
                after = after[:cut]

    before = _clean_context(before)
    after = _clean_context(after)
    if omitted_before and before:
        before = ELLIPSIS + before
    if omitted_after and after:
        after = after + ELLIPSIS
    return before, after


_BLOCK_MARKERS = re.compile(r"(?m)^[ \t]*(?:#{1,6}[ \t]+|>[ \t]?|[-*+][ \t]+|\d+[.)][ \t]+|```[^\n]*|~~~[^\n]*|\|)")
_TABLE_RULE = re.compile(r"(?m)^[ \t]*\|?[ \t]*:?-{3,}:?[ \t]*(?:\|[ \t]*:?-{3,}:?[ \t]*)*\|?[ \t]*$")
_LINK = re.compile(r"!?\[([^\]]*)\]\([^)]*\)")
_INLINE_MARKS = re.compile(r"[*`~]+|\|")
_EDGE_UNDERSCORE = re.compile(r"\b_+|_+\b")


def _clean_context(text: str) -> str:
    """Source markdown -> one line of readable prose (best effort, lossy)."""
    text = _TABLE_RULE.sub(" ", text)
    text = _BLOCK_MARKERS.sub("", text)
    text = _LINK.sub(r"\1", text)
    text = _INLINE_MARKS.sub("", text)
    text = _EDGE_UNDERSCORE.sub("", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


_NO_SPACE_BEFORE = ",.;:!?)]}…"
_NO_SPACE_AFTER = "([{"


def _join_context(before: str, span: str, after: str) -> str:
    """before + span + after with a space only where prose needs one (no
    space before ", but by" or after an opening bracket)."""
    out = before
    if before and span:
        out += "" if before[-1] in _NO_SPACE_AFTER else " "
    out += span
    if after:
        out += "" if after[0] in _NO_SPACE_BEFORE else " "
        out += after
    return out


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------

_MD_SPECIAL = re.compile(r"([\\`*_{}\[\]<>#|~])")
_MD_LINE_BULLET = re.compile(r"(?m)^([ \t]*)([-+])(?=[ \t])")
_MD_LINE_ORDERED = re.compile(r"(?m)^([ \t]*\d+)([.)])(?=[ \t])")


def escape_markdown(text: str) -> str:
    """Backslash-escape so ``text`` renders literally as Markdown text."""
    text = _MD_SPECIAL.sub(r"\\\1", text)
    text = _MD_LINE_BULLET.sub(r"\1\\\2", text)
    text = _MD_LINE_ORDERED.sub(r"\1\\\2", text)
    return text


def _fmt_date(dt: Optional[datetime]) -> str:
    return dt.strftime("%Y-%m-%d %H:%M") if dt else ""


def _blockquote(text: str) -> str:
    lines = escape_markdown(text).split("\n")
    return "\n".join(("> " + ln) if ln.strip() else ">" for ln in lines)


def _hard_break_lines(text: str) -> str:
    return "  \n".join(escape_markdown(ln) for ln in text.split("\n"))


def build_markdown(thread_title: str, rows: Sequence[ExportRow], now: Optional[datetime] = None) -> str:
    now = now or datetime.now()
    title = escape_markdown(thread_title or "Conversation")
    out: List[str] = [
        f"# Highlights — {title}",
        "",
        f"_{len(rows)} highlight{'' if len(rows) == 1 else 's'} · exported {now.strftime('%Y-%m-%d %H:%M')}_",
        "",
    ]
    if not rows:
        out += ["_No highlights in this thread._", ""]
        return "\n".join(out)
    for i, r in enumerate(rows, start=1):
        out += [f"## {i}.", "", _blockquote(r.selected_text), ""]
        if r.note.strip():
            out += [f"**Note:** {_hard_break_lines(r.note)}", ""]
        if r.context_available:
            span = f"**{escape_markdown(r.selected_text.replace(chr(10), ' '))}**"
            ctx = _join_context(escape_markdown(r.context_before), span, escape_markdown(r.context_after))
            out += [f"_Context:_ {ctx}", ""]
        else:
            out += ["_Context:_ (message no longer available)", ""]
        out += [f"_Colour: {escape_markdown(r.colour)} · {_fmt_date(r.created_at)}_", ""]
    return "\n".join(out)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _xlsx_context(r: ExportRow) -> str:
    if not r.context_available:
        return "(message no longer available)"
    return _join_context(r.context_before, "[" + r.selected_text.replace("\n", " ") + "]", r.context_after)


def build_xlsx(thread_title: str, rows: Sequence[ExportRow], now: Optional[datetime] = None) -> bytes:
    now = now or datetime.now()
    title = thread_title or "Conversation"
    table = {
        "name": "Highlights",
        "columns": [{"header": h} for h in XLSX_COLUMNS],
        "rows": [
            [_fmt_date(r.created_at), r.colour, r.selected_text, r.note, _xlsx_context(r), title]
            for r in rows
        ],
    }
    summary = {"Thread": title, "Highlights": len(rows)}
    raw = build_workbook({"tables": [table], "summary": summary}, now=now)

    # Post-pass (this module only): wrap the long text columns so multi-line
    # notes/text display as such, and give them a readable width.
    wb = load_workbook(io.BytesIO(raw))
    ws = wb["Highlights"]
    for col in XLSX_WRAP_COLUMNS:
        ws.column_dimensions[get_column_letter(col)].width = XLSX_WRAP_WIDTH
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Filenames
# ---------------------------------------------------------------------------

def export_filename(thread_title: str, fmt: str, now: Optional[datetime] = None) -> str:
    now = now or datetime.now()
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", thread_title or "").strip("_")[:40] or "thread"
    return f"highlights_{stem}_{now.strftime('%Y%m%d')}.{fmt}"
