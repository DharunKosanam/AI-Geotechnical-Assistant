"""Generic Excel (.xlsx) export of a calculator result.

The builder is GENERIC over the standard result schema (see
``app.workspace.calculators.base.ComputeResult``): it reads ``tables`` +
``summary`` off the persisted result object and never knows anything about CPT
(or any specific calculator). A NEW calculator gets the export for free by
declaring its tables/summary -- ZERO changes here.

Each table becomes one worksheet; a final "Summary" sheet is written from the
flat ``summary`` dict (plus a "Date generated" row stamped at export time).
Header rows are bold and frozen, numeric cells stay numeric with each column's
declared display format, columns are auto-sized, and there are no merged cells.

Only deterministic data is written -- the AI/interpretation text is NEVER
included. Plain ASCII throughout (Windows cp1252 console safe).
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True)
_MIN_WIDTH = 10
_MAX_WIDTH = 42
# Characters Excel forbids in a worksheet name, plus the 31-char length cap.
_SHEET_FORBIDDEN = re.compile(r"[:\\/?*\[\]]")


def export_filename(source_file: str, now: "datetime | None" = None) -> str:
    """Download filename like ``CPT_<sourcefilename>_<YYYYMMDD>.xlsx``.

    The source stem is sanitised to a safe ASCII token so the header value is
    always a valid, predictable filename.
    """
    now = now or datetime.now()
    stem = re.sub(r"\.[^.]*$", "", source_file or "sounding")  # drop extension
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("_") or "sounding"
    return f"CPT_{stem}_{now.strftime('%Y%m%d')}.xlsx"


def _safe_sheet_name(name: str, used: set) -> str:
    """A valid, unique (<=31 char) worksheet name."""
    clean = _SHEET_FORBIDDEN.sub("_", str(name or "Sheet")).strip() or "Sheet"
    clean = clean[:31]
    base, i = clean, 2
    while clean.lower() in used:
        suffix = f"_{i}"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


def _autosize(ws, widths: List[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            _MIN_WIDTH, min(_MAX_WIDTH, width + 2)
        )


def _is_number(value: Any) -> bool:
    # bool is an int subclass but must not be treated as a numeric cell.
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _write_table_sheet(ws, table: Dict[str, Any]) -> None:
    columns = table.get("columns", []) or []
    headers = [c.get("header", "") for c in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"  # keep the header row visible while scrolling

    widths = [len(h) for h in headers]
    for row in table.get("rows", []) or []:
        ws.append(list(row))
        for col_idx, col in enumerate(columns, start=1):
            value = row[col_idx - 1] if col_idx - 1 < len(row) else None
            fmt = col.get("format")
            if fmt and _is_number(value):
                ws.cell(row=ws.max_row, column=col_idx).number_format = fmt
            widths[col_idx - 1] = max(widths[col_idx - 1], len(str(value)))
    _autosize(ws, widths)


def _write_summary_sheet(ws, summary: Dict[str, Any], now: datetime) -> None:
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"

    rows = list(summary.items()) + [("Date generated", now.strftime("%Y-%m-%d"))]
    widths = [len("Field"), len("Value")]
    for label, value in rows:
        ws.append([label, value])
        widths[0] = max(widths[0], len(str(label)))
        widths[1] = max(widths[1], len(str(value)))
    _autosize(ws, widths)


def build_workbook(result_object: Dict[str, Any], now: "datetime | None" = None) -> bytes:
    """Serialize a result object's tables + summary to .xlsx bytes (generic).

    ``result_object`` is the persisted deterministic result: ``tables`` (each a
    ``{name, columns, rows}`` block) and a flat ``summary`` dict. Only that data
    is written -- never the AI text or the raw source file.
    """
    now = now or datetime.now()
    wb = Workbook()
    used_names: set = set()

    tables = result_object.get("tables", []) or []
    first = True
    for table in tables:
        ws = wb.active if first else wb.create_sheet()
        ws.title = _safe_sheet_name(table.get("name", "Data"), used_names)
        _write_table_sheet(ws, table)
        first = False

    summary_ws = wb.active if first else wb.create_sheet()
    summary_ws.title = _safe_sheet_name("Summary", used_names)
    _write_summary_sheet(summary_ws, result_object.get("summary", {}) or {}, now)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def result_has_tables(result_object: Dict[str, Any]) -> bool:
    """Whether a result declares at least one exportable table."""
    return bool(result_object.get("tables"))
