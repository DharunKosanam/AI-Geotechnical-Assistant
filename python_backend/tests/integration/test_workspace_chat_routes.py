"""Integration tests for the GeoPilot chat workspace routes.

Covers the session-document pool, the message->test chat routing, and the Excel
export. Exercised end-to-end through the ASGI app: auth is supplied by
overriding ``get_current_user``, the feature flag is toggled with monkeypatch,
and the LLM interpretation is mocked so nothing hits Ollama. The deterministic
parse -> calculator -> layer_summary chain and the openpyxl export run for real.
"""

import io
from datetime import datetime
from pathlib import Path

import pytest

from app.core import config
from app.dependencies.auth import get_current_user
from app.main import app
from app.workspace import store
from app.workspace.data import SAMPLE_CPT_FILE
from app.workspace.interpretation.ai_interpret import InterpretationResult
from models import User

pytestmark = pytest.mark.integration

_USER_ID = "507f1f77bcf86cd799439011"


def _fake_user() -> User:
    return User(
        id=_USER_ID,
        email="tester@example.com",
        hashed_password="x",
        full_name="Tester",
        created_at=datetime.now(),
        role="user",
    )


def _sample_bytes() -> bytes:
    return Path(SAMPLE_CPT_FILE).read_bytes()


@pytest.fixture
def authed():
    app.dependency_overrides[get_current_user] = _fake_user
    yield
    app.dependency_overrides.pop(get_current_user, None)


@pytest.fixture
def enable_workspace(monkeypatch):
    monkeypatch.setattr(config, "WORKSPACE_ENABLED", True)


@pytest.fixture(autouse=True)
def clean_store():
    store.clear_user(_USER_ID)
    yield
    store.clear_user(_USER_ID)


@pytest.fixture
def mock_interpret(monkeypatch):
    async def _fake(results, **kwargs):
        return InterpretationResult(
            narrative="Silt over sand over clay. AI draft for review.",
            per_layer_notes=[{"layer": 1}],
            flagged_concerns=["a concern"],
            model="test-model",
            is_ai_draft=True,
        )

    # Patch where the CPT plugin looks it up.
    import app.workspace.calculators.cpt as cpt_mod

    monkeypatch.setattr(cpt_mod, "interpret_sounding", _fake)


@pytest.fixture
def mock_router(monkeypatch):
    """Replace the LLM router client so non-trigger routing is hermetic.

    Mirrors the real router's job: pick cpt_interpretation when the message
    reads like a CPT request, else null. Records call count so a test can assert
    the fast-path skipped the LLM.
    """
    calls = {"count": 0}

    class _FakeRouterClient:
        async def chat(self, **kwargs):
            calls["count"] += 1
            prompt = kwargs["messages"][1]["content"].lower()
            # Inspect ONLY the actual user message, not the calculator catalog
            # (which itself contains "cpt"/"sounding"/"interpret").
            seg = prompt.split("user message:", 1)[-1].split("\n\n", 1)[0]
            cid = (
                "cpt_interpretation"
                if any(k in seg for k in ("cpt", "sounding", "interpret"))
                else None
            )
            import json

            return {"message": {"content": json.dumps({"calculator_id": cid})}}

    import app.workspace.router as router_mod

    monkeypatch.setattr(router_mod, "_default_client", lambda: _FakeRouterClient())
    return calls


# --- Documents -------------------------------------------------------------
@pytest.mark.asyncio
async def test_upload_list_delete_document(authed, enable_workspace, async_client):
    resp = await async_client.post(
        "/api/workspace/documents",
        files={"file": ("sample_sounding.CPT", _sample_bytes(), "text/plain")},
    )
    assert resp.status_code == 201, resp.text
    doc = resp.json()
    assert doc["filename"] == "sample_sounding.CPT"
    assert doc["extension"] == ".cpt"
    assert doc["status"] == "ready"

    listing = await async_client.get("/api/workspace/documents")
    assert listing.status_code == 200
    assert [d["id"] for d in listing.json()["documents"]] == [doc["id"]]

    deleted = await async_client.delete(f"/api/workspace/documents/{doc['id']}")
    assert deleted.status_code == 200
    assert deleted.json() == {"deleted": doc["id"]}

    missing = await async_client.delete(f"/api/workspace/documents/{doc['id']}")
    assert missing.status_code == 404


@pytest.mark.asyncio
async def test_upload_empty_file_rejected(authed, enable_workspace, async_client):
    resp = await async_client.post(
        "/api/workspace/documents",
        files={"file": ("empty.cpt", b"", "text/plain")},
    )
    assert resp.status_code == 400


@pytest.mark.asyncio
async def test_documents_gated_by_flag(authed, monkeypatch, async_client):
    # With WORKSPACE_ENABLED off, the route behaves as if it does not exist.
    monkeypatch.setattr(config, "WORKSPACE_ENABLED", False)
    resp = await async_client.get("/api/workspace/documents")
    assert resp.status_code == 404


# --- Chat routing (message -> test) ----------------------------------------
async def _upload_sample(async_client) -> str:
    resp = await async_client.post(
        "/api/workspace/documents",
        files={"file": ("sample_sounding.CPT", _sample_bytes(), "text/plain")},
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


@pytest.mark.asyncio
async def test_chat_unknown_message_lists_tests(
    authed, enable_workspace, mock_router, async_client
):
    resp = await async_client.post("/api/workspace/chat", json={"message": "hello"})
    assert resp.status_code == 200
    data = resp.json()
    assert data["type"] == "info"
    assert "CPT interpretation" in data["text"]


@pytest.mark.asyncio
async def test_chat_unregistered_test_is_graceful(
    authed, enable_workspace, mock_router, async_client
):
    resp = await async_client.post(
        "/api/workspace/chat", json={"message": "run terzaghi"}
    )
    assert resp.status_code == 200
    # 'run terzaghi' is not a registered calculator -> router null -> lists tests.
    assert resp.json()["type"] == "info"


@pytest.mark.asyncio
async def test_chat_natural_phrasing_routes_to_cpt(
    authed, enable_workspace, mock_router, mock_interpret, async_client
):
    # Phrasing with no exact trigger phrase must route via the LLM router.
    await _upload_sample(async_client)
    resp = await async_client.post(
        "/api/workspace/chat", json={"message": "can you interpret this sounding?"}
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["type"] == "result"
    assert mock_router["count"] == 1  # LLM was consulted


@pytest.mark.asyncio
async def test_chat_exact_phrase_uses_fast_path(
    authed, enable_workspace, mock_router, mock_interpret, async_client
):
    # 'run cpt' is an exact trigger -> fast-path, the router LLM must NOT fire.
    await _upload_sample(async_client)
    resp = await async_client.post("/api/workspace/chat", json={"message": "run cpt"})
    assert resp.status_code == 200, resp.text
    assert resp.json()["type"] == "result"
    assert mock_router["count"] == 0  # fast-path skipped the LLM


@pytest.mark.asyncio
async def test_chat_trigger_without_doc_asks_for_upload(
    authed, enable_workspace, async_client
):
    resp = await async_client.post("/api/workspace/chat", json={"message": "run CPT"})
    assert resp.status_code == 200
    data = resp.json()
    assert data["type"] == "need_upload"
    assert "upload" in data["text"].lower()


@pytest.mark.asyncio
async def test_chat_runs_cpt_with_doc(
    authed, enable_workspace, mock_interpret, async_client
):
    await _upload_sample(async_client)
    resp = await async_client.post("/api/workspace/chat", json={"message": "run CPT"})
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["type"] == "result"
    assert data["calculator_id"] == "cpt_interpretation"
    assert data["source_file"] == "sample_sounding.CPT"
    assert data["layers"], "expected detected layers in the result"
    assert data["interpretation"]["narrative"].startswith("Silt over sand")
    assert data["reference"]
    assert data["run_id"]  # persisted -> durable export id
    assert data["exportable"] is True
    assert data["thread_id"]  # turn recorded in a thread


@pytest.mark.asyncio
async def test_chat_parses_inline_params(
    authed, enable_workspace, mock_interpret, async_client
):
    await _upload_sample(async_client)
    resp = await async_client.post(
        "/api/workspace/chat",
        json={"message": "run cpt, groundwater 2m, unit weight 18"},
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["params"] == {"groundwater_level": 2.0, "soil_unit_weight": 18.0}
    assert data["metadata"]["groundwater_level"] == 2.0
    assert data["metadata"]["soil_unit_weight"] == 18.0


# --- Excel export ----------------------------------------------------------
@pytest.mark.asyncio
async def test_export_xlsx_download_survives_lost_memory(
    authed, enable_workspace, mock_interpret, async_client
):
    await _upload_sample(async_client)
    run = await async_client.post("/api/workspace/chat", json={"message": "run CPT"})
    run_id = run.json()["run_id"]
    assert run_id, "run should be persisted and expose a run_id"

    # Simulate a restart: drop ALL in-memory session state. Export must still
    # work because it reads the persisted run from the (fake) DB.
    store.clear_user(_USER_ID)

    resp = await async_client.get(f"/api/workspace/export/{run_id}")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disp = resp.headers["content-disposition"]
    assert "CPT_sample_sounding_" in disp and disp.endswith('.xlsx"')

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["CPT Data", "Summary"]


@pytest.mark.asyncio
async def test_export_unknown_run_is_clean_404(authed, enable_workspace, async_client):
    # Malformed and valid-but-missing ids both -> clean 404 with the re-run hint.
    for bad in ["does-not-exist", str(__import__("bson").ObjectId())]:
        resp = await async_client.get(f"/api/workspace/export/{bad}")
        assert resp.status_code == 404
        assert "re-run" in resp.json()["detail"].lower()


# --- Scoped Q&A (non-command messages) -------------------------------------
@pytest.fixture
def mock_qa(monkeypatch):
    """Replace the scoped-Q&A LLM client with a canned answer; capture the prompt."""

    captured = {}

    class _FakeClient:
        async def chat(self, **kwargs):
            captured["messages"] = kwargs["messages"]
            return {"message": {"content": "Layer 3 (clay) has low qc - a concern. AI draft."}}

    import app.workspace.interpretation.qa as qa_mod

    monkeypatch.setattr(qa_mod, "_default_client", lambda: _FakeClient())
    return captured


@pytest.mark.asyncio
async def test_question_before_any_result_lists_tests(
    authed, enable_workspace, mock_router, async_client
):
    resp = await async_client.post(
        "/api/workspace/chat", json={"message": "is layer 3 a concern?"}
    )
    assert resp.status_code == 200
    # No result yet -> still the info/list-tests behaviour, no Q&A fired.
    assert resp.json()["type"] == "info"


@pytest.mark.asyncio
async def test_question_after_result_is_scoped_answer(
    authed, enable_workspace, mock_router, mock_interpret, mock_qa, async_client
):
    await _upload_sample(async_client)
    await async_client.post("/api/workspace/chat", json={"message": "run CPT"})

    resp = await async_client.post(
        "/api/workspace/chat", json={"message": "is layer 3 a concern?"}
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["type"] == "answer"
    assert data["is_ai_draft"] is True
    assert data["answer"].startswith("Layer 3 (clay)")

    # The scoped result context (not RAG) was handed to the model.
    sys_msg = mock_qa["messages"][0]["content"]
    user_msg = mock_qa["messages"][1]["content"]
    assert "strictly about THIS CPT sounding result" in sys_msg
    assert "LAYERS (top to bottom):" in user_msg
    assert "is layer 3 a concern?" in user_msg


@pytest.mark.asyncio
async def test_registered_trigger_still_runs_after_result(
    authed, enable_workspace, mock_interpret, mock_qa, async_client
):
    # A trigger message must still route to the calculator, not the Q&A path.
    await _upload_sample(async_client)
    await async_client.post("/api/workspace/chat", json={"message": "run CPT"})
    resp = await async_client.post("/api/workspace/chat", json={"message": "run CPT"})
    assert resp.json()["type"] == "result"
