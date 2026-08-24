"""Integration tests for the instrument-dataset path (INSTRUMENT_PARSERS_ENABLED).

Exercises the FastAPI app via ASGI: the sniff branch in the GeoPilot upload
route, the background parse job (runs inside the ASGI call), the dataset /
job endpoints, delete + retry, and -- most importantly -- flag-off parity:
with the flag off nothing is sniffed, the status payload is exactly the
pre-feature one and the dataset routes do not exist (404).

Mongo is replaced by in-memory fakes; the arrays land in a tmp
INSTRUMENT_DATA_DIR so nothing touches the real data dir.
"""

from __future__ import annotations

import copy
import os
from datetime import datetime
from types import SimpleNamespace

import numpy as np
import pytest

from app.core import config
from app.dependencies.auth import get_current_user
from app.main import app
from app.workspace import dataset_routes, dataset_store, store
from models import User
from tests.integration.conftest import FakeCollection

pytestmark = pytest.mark.integration


class _Coll(FakeCollection):
    """FakeCollection + the delete verbs the dataset store uses."""

    async def delete_one(self, query):
        for _id, doc in list(self._docs.items()):
            if self._match(doc, query):
                del self._docs[_id]
                return SimpleNamespace(deleted_count=1)
        return SimpleNamespace(deleted_count=0)

    async def delete_many(self, query):
        n = 0
        for _id, doc in list(self._docs.items()):
            if self._match(doc, query):
                del self._docs[_id]
                n += 1
        return SimpleNamespace(deleted_count=n)


USER_ID = "507f1f77bcf86cd799439011"


def _user(uid: str = USER_ID) -> User:
    return User(
        id=uid, email="tester@example.com", hashed_password="x", full_name="Tester",
        created_at=datetime.now(), role="user",
    )


@pytest.fixture
def authed():
    app.dependency_overrides[get_current_user] = _user
    yield
    app.dependency_overrides.pop(get_current_user, None)
    store.clear_user(USER_ID)


@pytest.fixture
def workspace_on(monkeypatch):
    monkeypatch.setattr(config, "WORKSPACE_ENABLED", True)


@pytest.fixture
def flag_on(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "INSTRUMENT_PARSERS_ENABLED", True)
    monkeypatch.setattr(config, "INSTRUMENT_DATA_DIR", str(tmp_path / "instrument_datasets"))


@pytest.fixture
def flag_off(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "INSTRUMENT_PARSERS_ENABLED", False)
    monkeypatch.setattr(config, "INSTRUMENT_DATA_DIR", str(tmp_path / "instrument_datasets"))


@pytest.fixture(autouse=True)
def register_dataset_routes(monkeypatch):
    """The dataset router is included at startup only when the flag is on
    (route table parity). Register it once for this module so the call-time
    gate is what decides 404 vs served in each test."""
    monkeypatch.setattr(config, "INSTRUMENT_PARSERS_ENABLED", True)
    dataset_routes.register(app)


@pytest.fixture(autouse=True)
def dataset_db(monkeypatch):
    ds, jobs = _Coll(), _Coll()
    monkeypatch.setattr(dataset_store, "datasets_collection", ds)
    monkeypatch.setattr(dataset_store, "jobs_collection", jobs)
    return SimpleNamespace(datasets=ds, jobs=jobs)


def _odisi_bytes(n_gages: int = 6, n_steps: int = 4) -> bytes:
    header = [
        "Test Name:\tunit", "Product:\tODiSI 6104", "Sensor Serial Number:\tFS02025LUNA0017736",
        "Sensor Length (m):\t20.4542", "Gage Pitch (mm):\t2.6",
        "Measurement Rate per Channel (Hz):\t8.333", "Tare Name:\t0409", "Units:\tmicrostrain",
    ]
    x = [0.08 + 0.0026 * i for i in range(n_gages)]
    lines = header + ["-" * 20]
    lines.append("Tare\t\t\t" + "\t".join(f"{1.0 + i:.3f}" for i in range(n_gages)))
    lines.append("x-axis\t\t\t" + "\t".join(f"{v:.4f}" for v in x))
    for k in range(n_steps):
        vals = "\t".join(f"{10 * k + i:.3f}" for i in range(n_gages))
        lines.append(f"2026-04-09 16:23:{46 + k:02d}.000\tmeasurement\t\t" + vals)
    return ("\n".join(lines) + "\n").encode()


def _campbell_bytes(n: int = 30) -> bytes:
    lines = ["TIMESTAMP,RECORD,TP4144_kPa,TP4145_kPa"]
    for i in range(n):
        lines.append(f"2024-05-10 00:00:{i // 10:02d}.{(i % 10) * 100:03d},{249671 + i},{11.6 + 0.01 * i:.3f},{9.9:.3f}")
    return ("\n".join(lines) + "\n").encode()


def _pdf_bytes() -> bytes:
    return b"%PDF-1.4\n1 0 obj\n<< /Type /Catalog >>\nendobj\n%%EOF\n"


async def _upload(client, name: str, data: bytes):
    return await client.post(
        "/api/workspace/documents", files={"file": (name, data, "application/octet-stream")}
    )


# --- Flag OFF: byte-identical to before ------------------------------------
async def test_flag_off_status_payload_is_exactly_pre_feature(async_client, authed, workspace_on, flag_off):
    res = await async_client.get("/api/workspace/status")
    assert res.status_code == 200
    assert res.json() == {"enabled": True}
    assert res.content == b'{"enabled":true}'


async def test_flag_off_dataset_routes_do_not_exist(async_client, authed, workspace_on, flag_off):
    for method, url in (
        ("GET", "/api/workspace/datasets"),
        ("GET", "/api/workspace/datasets/507f1f77bcf86cd799439011"),
        ("GET", "/api/workspace/datasets/jobs/507f1f77bcf86cd799439011"),
        ("DELETE", "/api/workspace/datasets/507f1f77bcf86cd799439011"),
        ("POST", "/api/workspace/datasets/507f1f77bcf86cd799439011/retry"),
    ):
        res = await async_client.request(method, url)
        assert res.status_code == 404, (method, url, res.status_code)
        assert res.json() == {"detail": "Not Found"}  # exactly like an absent route


async def test_flag_off_tsv_and_pdf_take_document_path_unchanged(async_client, authed, workspace_on, flag_off, dataset_db):
    tsv = _odisi_bytes()
    res = await _upload(async_client, "pass_001.tsv", tsv)
    assert res.status_code == 201
    body = res.json()
    assert set(body) == {"id", "filename", "extension", "status"}  # pre-feature shape
    assert body["status"] == "ready" and body["extension"] == ".tsv"
    doc = store.get_document(USER_ID, body["id"])
    assert doc is not None and doc.text == tsv.decode()  # whole file, as text, as before

    res = await _upload(async_client, "doc.pdf", _pdf_bytes())
    assert res.status_code == 201 and set(res.json()) == {"id", "filename", "extension", "status"}
    # No dataset artifact, no job, nothing on disk.
    assert dataset_db.datasets._docs == {} and dataset_db.jobs._docs == {}
    assert not os.path.exists(config.INSTRUMENT_DATA_DIR)


# --- Flag ON: sniff -> parser path -------------------------------------------
async def test_flag_on_status_advertises_capability(async_client, authed, workspace_on, flag_on):
    res = await async_client.get("/api/workspace/status")
    body = res.json()
    assert body["enabled"] is True and body["instrument_parsers"] is True
    assert ".tsv" in body["instrument_extensions"] and ".dat" in body["instrument_extensions"]


async def test_flag_on_odisi_upload_parses_to_artifact_without_arrays_in_mongo(
    async_client, authed, workspace_on, flag_on, dataset_db
):
    res = await _upload(async_client, "ODiSI_pass_001.tsv", _odisi_bytes(n_gages=6, n_steps=4))
    assert res.status_code == 201
    body = res.json()
    assert body["kind"] == "dataset" and body["status"] == "queued"
    assert body["parser_id"] == "odisi_tsv" and body["dataset_kind"] == "strain_distributed"
    ds_id, job_id = body["dataset_id"], body["job_id"]
    # Not in the document store (never decoded as text).
    assert store.list_documents(USER_ID) == []

    # The background task ran inside the ASGI call: dataset is parsed.
    res = await async_client.get("/api/workspace/datasets")
    rows = res.json()["datasets"]
    assert len(rows) == 1
    row = rows[0]
    assert row["id"] == ds_id and row["status"] == "parsed" and row["progress"] == 100
    assert row["badge"] == "DFOS · 6 gages"
    assert row["metadata"]["n_gages"] == 6 and row["metadata"]["n_timesteps"] == 4
    assert row["metadata"]["tare_name"] == "0409"
    assert row["metadata"]["source_filename"] == "ODiSI_pass_001.tsv"  # original name, not the on-disk id
    assert "_raw_header" not in row["metadata"]  # list view is compact
    assert row["shapes"] == {"x_axis": [6], "tare": [6], "strain": [4, 6], "timestamps": [4], "timestamp_text": [4]}
    assert row["dtypes"]["strain"] == "float32"
    assert row["warnings"] == [] and row["segments"] == []

    res = await async_client.get(f"/api/workspace/datasets/{ds_id}")
    full = res.json()
    assert full["metadata"]["_raw_header"][0] == "Test Name:\tunit"
    assert full["npz_path"] and os.path.exists(full["npz_path"])
    assert full["raw_path"] and os.path.exists(full["raw_path"])
    with np.load(full["npz_path"]) as npz:
        assert npz["strain"].shape == (4, 6) and npz["strain"].dtype == np.float32
        assert float(npz["x_axis"][0]) == pytest.approx(0.08)

    res = await async_client.get(f"/api/workspace/datasets/jobs/{job_id}")
    job = res.json()
    assert job["state"] == "parsed" and job["progress"] == 100 and job["error"] is None
    assert job["dataset_id"] == ds_id and job["elapsed_s"] is not None

    # The Mongo pointer doc holds NO array data.
    stored = next(iter(dataset_db.datasets._docs.values()))
    for key, value in stored.items():
        assert not isinstance(value, np.ndarray), key
    assert "arrays" not in stored and "strain" not in stored
    assert stored["npz_path"].endswith(f"{ds_id}.npz")


async def test_flag_on_campbell_upload_and_pdf_still_documents(async_client, authed, workspace_on, flag_on, dataset_db):
    res = await _upload(async_client, "2024-05-10.dat", _campbell_bytes(30))
    assert res.status_code == 201 and res.json()["kind"] == "dataset"
    res = await async_client.get("/api/workspace/datasets")
    row = res.json()["datasets"][0]
    assert row["status"] == "parsed" and row["badge"] == "Pressure · 2 channels"
    assert row["metadata"]["channel_names"] == ["TP4144_kPa", "TP4145_kPa"]
    assert row["metadata"]["n_samples"] == 30 and row["metadata"]["sample_rate_hz"] == 10.0

    # A PDF (no signature) still takes the document path, byte-for-byte.
    res = await _upload(async_client, "doc.pdf", _pdf_bytes())
    assert res.status_code == 201
    assert set(res.json()) == {"id", "filename", "extension", "status"}
    doc = store.get_document(USER_ID, res.json()["id"])
    assert doc.text == _pdf_bytes().decode("utf-8", errors="replace")
    assert len(dataset_db.datasets._docs) == 1  # the PDF created no artifact


async def test_flag_on_large_document_is_rejoined_after_sniff(async_client, authed, workspace_on, flag_on):
    # A non-instrument text file bigger than the 2 KB sniff window must arrive
    # in the document store WHOLE (head + rest re-joined).
    text = ("$\nHA=1,MA=0.80\n#\n" + "\n".join(f"D={i / 100:.2f},QC=1.0,FS=10,U=5" for i in range(400)) + "\n")
    assert len(text) > 4096
    res = await _upload(async_client, "sounding.CPT", text.encode())
    assert res.status_code == 201 and "kind" not in res.json()
    assert store.get_document(USER_ID, res.json()["id"]).text == text


async def test_failed_parse_is_reported_and_retry_requeues(async_client, authed, workspace_on, flag_on, dataset_db):
    # Header-only Campbell file: sniffs as Campbell, parser raises ParserError.
    res = await _upload(async_client, "empty.dat", b"TIMESTAMP,RECORD,TP4144_kPa\n")
    assert res.status_code == 201
    ds_id = res.json()["dataset_id"]
    row = (await async_client.get("/api/workspace/datasets")).json()["datasets"][0]
    assert row["status"] == "failed" and "no data rows" in row["error"]
    old_job = row["job_id"]

    res = await async_client.post(f"/api/workspace/datasets/{ds_id}/retry")
    assert res.status_code == 200
    row = (await async_client.get("/api/workspace/datasets")).json()["datasets"][0]
    assert row["job_id"] != old_job  # a fresh job was created ...
    assert row["status"] == "failed"  # ... and honestly failed again on the same file


async def test_delete_removes_pointer_jobs_and_files(async_client, authed, workspace_on, flag_on, dataset_db):
    res = await _upload(async_client, "pass.tsv", _odisi_bytes())
    ds_id = res.json()["dataset_id"]
    full = (await async_client.get(f"/api/workspace/datasets/{ds_id}")).json()
    assert os.path.exists(full["npz_path"]) and os.path.exists(full["raw_path"])
    res = await async_client.delete(f"/api/workspace/datasets/{ds_id}")
    assert res.status_code == 200 and res.json() == {"deleted": ds_id}
    assert not os.path.exists(full["npz_path"]) and not os.path.exists(full["raw_path"])
    assert dataset_db.datasets._docs == {} and dataset_db.jobs._docs == {}
    assert (await async_client.get(f"/api/workspace/datasets/{ds_id}")).status_code == 404


async def test_datasets_are_user_scoped(async_client, authed, workspace_on, flag_on, dataset_db):
    res = await _upload(async_client, "pass.tsv", _odisi_bytes())
    ds_id, job_id = res.json()["dataset_id"], res.json()["job_id"]
    app.dependency_overrides[get_current_user] = lambda: _user("507f1f77bcf86cd799439099")
    try:
        assert (await async_client.get("/api/workspace/datasets")).json() == {"datasets": []}
        assert (await async_client.get(f"/api/workspace/datasets/{ds_id}")).status_code == 404
        assert (await async_client.get(f"/api/workspace/datasets/jobs/{job_id}")).status_code == 404
        assert (await async_client.delete(f"/api/workspace/datasets/{ds_id}")).status_code == 404
        assert (await async_client.post(f"/api/workspace/datasets/{ds_id}/retry")).status_code == 404
    finally:
        app.dependency_overrides[get_current_user] = _user
    assert len(dataset_db.datasets._docs) == 1


async def test_bad_ids_are_clean_404s(async_client, authed, workspace_on, flag_on):
    assert (await async_client.get("/api/workspace/datasets/not-an-id")).status_code == 404
    assert (await async_client.get("/api/workspace/datasets/jobs/zzz")).status_code == 404


# --- read-time staleness + registration guard (pure) ---------------------------
def test_effective_state_reports_interrupted_parse_after_timeout(monkeypatch):
    from datetime import timedelta, timezone

    monkeypatch.setattr(config, "INSTRUMENT_PARSE_TIMEOUT_SECONDS", 60)
    old = datetime.now(timezone.utc) - timedelta(seconds=120)
    state, err = dataset_store.effective_state({"status": "parsing", "progress": 40, "updated_at": old})
    assert state == "failed" and "interrupted" in err
    fresh = datetime.now(timezone.utc)
    assert dataset_store.effective_state({"status": "parsing", "updated_at": fresh}) == ("parsing", None)
    assert dataset_store.effective_state({"status": "parsed", "updated_at": old}) == ("parsed", None)
    assert dataset_store.effective_state({"status": "failed", "error": "x", "updated_at": old}) == ("failed", "x")


def test_register_adds_nothing_when_flag_off(monkeypatch):
    from fastapi import FastAPI

    monkeypatch.setattr(config, "INSTRUMENT_PARSERS_ENABLED", False)
    monkeypatch.setattr(dataset_routes, "_registered", False)
    fresh_app = FastAPI()
    before = len(fresh_app.routes)
    dataset_routes.register(fresh_app)
    assert len(fresh_app.routes) == before
    monkeypatch.setattr(config, "INSTRUMENT_PARSERS_ENABLED", True)
    dataset_routes.register(fresh_app)
    assert len(fresh_app.routes) > before
    monkeypatch.setattr(dataset_routes, "_registered", True)  # leave the shared app registered


# --- Dataset-bound calculator through the chat route ---------------------------
class _FakeOllama:
    def __init__(self):
        self.calls = []

    async def chat(self, **kwargs):
        self.calls.append(kwargs)
        return {"message": {"content": "Draft: peak strain noted. AI draft for review."}}


@pytest.fixture
def fake_llm(monkeypatch):
    fake = _FakeOllama()
    monkeypatch.setattr("app.workspace.interpretation.dataset_interpret._default_client", lambda: fake)
    return fake


@pytest.fixture(autouse=True)
def hermetic_router(monkeypatch):
    """Never reach the live LLM router: the fast exact-phrase pre-check decides,
    anything else routes to null (-> info / scoped Q&A branch)."""
    import json

    class _NullRouter:
        async def chat(self, **kwargs):
            return {"message": {"content": json.dumps({"calculator_id": None})}}

    import app.workspace.router as router_mod

    monkeypatch.setattr(router_mod, "_default_client", lambda: _NullRouter())


async def _run_chat(client, message: str):
    return await client.post("/api/workspace/chat", json={"message": message})


async def test_dfos_calculator_needs_a_dataset_then_runs_deterministically(
    async_client, authed, workspace_on, flag_on, dataset_db, fake_llm, monkeypatch
):
    # No dataset yet -> need_upload naming the Datasets panel.
    res = await _run_chat(async_client, "run dfos pass strain")
    assert res.status_code == 200
    body = res.json()
    assert body["type"] == "need_upload" and body["calculator_id"] == "dfos_pass_strain"
    assert "Datasets panel" in body["text"]

    # Upload + parse a small ODiSI file, then run.
    up = await _upload(async_client, "ODiSI_pass_001.tsv", _odisi_bytes(n_gages=40, n_steps=12))
    ds_id = up.json()["dataset_id"]
    # 40 gages span 0.1 m: the default 1.10 m lead-in + 0.50 m tail exclusions
    # leave nothing -> a graceful, explicit error (not a 500).
    res = await _run_chat(async_client, "run dfos pass strain")
    assert res.json()["type"] == "error" and "leave fewer than 3 of 40 gages" in res.json()["text"]
    monkeypatch.setattr(config, "DFOS_LEADIN_EXCLUDE_M", 0.09)  # x starts at 0.08 -> a few gages
    monkeypatch.setattr(config, "DFOS_TAIL_EXCLUDE_M", 0.01)
    res = await _run_chat(async_client, "run dfos pass strain, influence line at 0.1 m")
    body = res.json()
    assert body["type"] == "result", body
    assert body["calculator_id"] == "dfos_pass_strain" and body["dataset_id"] == ds_id
    assert body["source_file"] == "ODiSI_pass_001.tsv"
    assert body["params"] == {"gage_x": 0.1}
    assert [c["id"] for c in body["charts"]] == ["envelope", "band_profile", "influence", "load_position"]
    assert all(len(s["x"]) <= 2000 for c in body["charts"] for s in c["series"])
    assert body["notices"][0]["level"] == "provisional"  # exclusion notice, visible in the block
    assert "lead-in 0.09 m" in body["notices"][0]["text"] and "tail 0.01 m" in body["notices"][0]["text"]
    assert body["summary"]["Lead-in excluded at head (m) - unbonded lead-in (provisional)"] == 0.09
    assert body["summary"]["Tail excluded at far end (m) - fibre termination artifact (provisional)"] == 0.01
    assert "Analysed span (m)" in body["summary"] and "NaN fraction (%)" in body["summary"]
    assert "Peak tracking" in body["summary"]
    assert body["charts"][0]["bands"][0]["label"] == "lead-in 0.09 m excluded"
    assert any(c["id"] == "band_profile" for c in body["charts"])
    assert body["metadata"]["band_profile"]
    assert body["layers"] == []
    assert body["interpretation"]["is_ai_draft"] is True
    assert body["interpretation"]["narrative"].startswith("Draft:")
    assert body["exportable"] is True and body["run_id"]
    assert fake_llm.calls and fake_llm.calls[0]["think"] is True

    # Determinism: run again -> identical deterministic payload (AI text aside).
    res2 = await _run_chat(async_client, "run dfos pass strain, influence line at 0.1 m")
    b2 = res2.json()
    for key in ("summary_text", "summary", "metadata", "charts", "notices"):
        assert b2[key] == body[key], key

    # Export by run id: workbook has the three tables + Summary.
    import io
    from openpyxl import load_workbook

    ex = await async_client.get(f"/api/workspace/export/{body['run_id']}")
    assert ex.status_code == 200
    assert 'filename="DFOS_ODiSI_pass_001_' in ex.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(ex.content))
    assert wb.sheetnames == ["Peak per timestep", "Envelope", "Influence line", "Band profile", "Summary"]
    assert wb["Envelope"].max_row == 41  # header + 40 gages, full precision


async def test_dataset_selection_by_filename_else_newest(async_client, authed, workspace_on, flag_on, fake_llm, monkeypatch):
    monkeypatch.setattr(config, "DFOS_LEADIN_EXCLUDE_M", 0.085)
    monkeypatch.setattr(config, "DFOS_TAIL_EXCLUDE_M", 0.005)
    await _upload(async_client, "pass_001.tsv", _odisi_bytes(n_gages=20, n_steps=6))
    await _upload(async_client, "pass_002.tsv", _odisi_bytes(n_gages=20, n_steps=8))
    newest = (await _run_chat(async_client, "run dfos pass strain")).json()
    assert newest["source_file"] == "pass_002.tsv"
    named = (await _run_chat(async_client, "run dfos pass strain on pass 001")).json()
    assert named["source_file"] == "pass_001.tsv"


async def test_flag_off_hides_dataset_calculators_from_the_catalog(async_client, authed, workspace_on, flag_off):
    res = await _run_chat(async_client, "hello there")
    assert res.json()["type"] == "info"
    assert "DFOS" not in res.json()["text"]
    assert res.json()["text"] == "I can run: CPT interpretation - say 'run cpt'."
    # And the trigger phrase does not run anything with the flag off.
    res = await _run_chat(async_client, "run dfos pass strain")
    assert res.json()["type"] == "info"


def _campbell_events_bytes(n_events: int = 6, minutes: int = 20) -> bytes:
    """A pressure log with clear synthetic passes (4 channels in a line)."""
    import numpy as np

    n = minutes * 60 * 10
    t = np.arange(n) * 0.1
    rng = np.random.default_rng(7)
    base = np.array([11.6, 9.9, 10.2, 7.5])
    p = base[None, :] + rng.normal(0, 0.1, (n, 4))
    for k, tc in enumerate(np.linspace(60, minutes * 60 - 60, n_events)):
        for c in range(4):
            p[:, c] += (12 + 2 * k) * np.exp(-((t - tc - c * 0.4) / 0.5) ** 2)
    lines = ["TIMESTAMP,RECORD,TP4144_kPa,TP4145_kPa,TP4148_kPa,TP4149_kPa"]
    for i in range(n):
        ms = int(round(t[i] * 1000))
        ts = f"2024-05-10 00:{(ms // 60000) % 60:02d}:{(ms // 1000) % 60:02d}.{ms % 1000:03d}"
        lines.append(f"{ts},{249671 + i}," + ",".join(f"{v:.3f}" for v in p[i]))
    return ("\n".join(lines) + "\n").encode()


async def test_traffic_calculator_is_provisional_and_attaches_segments(
    async_client, authed, workspace_on, flag_on, dataset_db, fake_llm
):
    up = await _upload(async_client, "2024-05-10.dat", _campbell_events_bytes(6))
    ds_id = up.json()["dataset_id"]
    res = await _run_chat(async_client, "run traffic load monitoring")
    body = res.json()
    assert body["type"] == "result" and body["calculator_id"] == "traffic_load_monitoring"
    assert body["dataset_id"] == ds_id
    # Provisional status is first-class in the reply (renders in the card block).
    assert body["notices"][0]["level"] == "provisional"
    assert "pending validation by the supervising engineer" in body["notices"][0]["text"]
    assert body["summary"]["Validation status"].startswith("PROVISIONAL")
    assert "PROVISIONAL" in body["reference"]
    assert body["metadata"]["n_events"] == 6
    assert len(body["segments"]) == 6
    assert body["segments"][0]["direction"] == "TP4144_kPa → TP4149_kPa"
    assert body["exportable"] is True

    # Segments became children of the dataset row.
    ds = (await async_client.get(f"/api/workspace/datasets/{ds_id}")).json()
    assert len(ds["segments"]) == 6 and ds["segments"][0]["label"].startswith("Event 1")
    listed = (await async_client.get("/api/workspace/datasets")).json()["datasets"][0]
    assert len(listed["segments"]) == 6

    # Determinism.
    b2 = (await _run_chat(async_client, "run traffic load monitoring")).json()
    for key in ("summary_text", "summary", "metadata", "charts", "notices", "segments"):
        assert b2[key] == body[key], key

    # Inline parameter override is honoured and echoed.
    b3 = (await _run_chat(async_client, "run traffic load monitoring, mad multiplier 200")).json()
    assert b3["params"] == {"mad_multiplier": 200.0}
    assert b3["metadata"]["detection"]["mad_multiplier"] == 200.0

    # Export carries the provisional status.
    import io
    from openpyxl import load_workbook

    ex = await async_client.get(f"/api/workspace/export/{body['run_id']}")
    assert ex.status_code == 200
    assert 'filename="TRAFFIC_2024-05-10_' in ex.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(ex.content))
    assert wb.sheetnames[0] == "Events (provisional)"
    summ = {row[0].value: row[1].value for row in wb["Summary"].iter_rows(min_row=2)}
    assert summ["Validation status"].startswith("PROVISIONAL")
