"""Unit tests for the GENERIC Excel export + CPT parity.

The builder is generic over the standard result schema (tables + summary). These
tests confirm: (a) a CPT result exports two sheets with numeric cells, bold/
frozen headers and no AI text, and (b) the refactor is byte-for-content
identical to the pre-refactor CPT output (golden snapshots under data/).
"""

import io
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.workspace.calculators.cpt import CPT_CALCULATOR
from app.workspace.data import SAMPLE_CPT_FILE
from app.workspace.export.xlsx import (
    build_workbook,
    export_filename,
    result_has_tables,
)

pytestmark = pytest.mark.unit

_DATA_DIR = Path(__file__).resolve().parent / "data"
# Fixed date so the "Date generated" Summary row is deterministic (matches golden).
_NOW = datetime(2026, 7, 14)


def _result_object(params=None):
    text = Path(SAMPLE_CPT_FILE).read_text(encoding="utf-8", errors="replace")
    result = CPT_CALCULATOR.compute(text, "sample_sounding.CPT", params or {})
    return {
        "calculator_id": "cpt_interpretation",
        "calculator_name": CPT_CALCULATOR.name,
        "source_file": "sample_sounding.CPT",
        "reference": CPT_CALCULATOR.reference,
        "layers": result.layers,
        "metadata": result.metadata,
        "tables": result.tables,
        "summary": result.summary,
    }


def _dump(xlsx_bytes: bytes) -> dict:
    """Read a workbook into a comparable content structure (values + formats)."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    out = {}
    for ws in wb.worksheets:
        rows, formats = [], []
        for row in ws.iter_rows():
            rows.append([c.value for c in row])
            formats.append([c.number_format for c in row])
        out[ws.title] = {
            "rows": rows,
            "formats": formats,
            "freeze": ws.freeze_panes,
            "header_bold": [c.font.bold for c in ws[1]],
            "widths": [
                ws.column_dimensions[chr(ord("A") + i)].width
                for i in range(ws.max_column)
            ],
        }
    return out


# --- Generic behaviour -----------------------------------------------------
def test_export_filename_shape():
    assert export_filename("my sounding.CPT", now=_NOW) == "CPT_my_sounding_20260714.xlsx"


def test_result_has_tables():
    assert result_has_tables(_result_object()) is True
    assert result_has_tables({"tables": []}) is False
    assert result_has_tables({}) is False


def test_two_sheets_numeric_and_no_ai_text():
    payload = _result_object()
    payload["ai_narrative"] = "SHOULD NOT APPEAR silt over sand"  # must be ignored
    wb = load_workbook(io.BytesIO(build_workbook(payload, now=_NOW)))
    assert wb.sheetnames == ["CPT Data", "Summary"]

    ws = wb["CPT Data"]
    assert ws["A1"].value == "Depth (m)"
    assert ws["A1"].font.bold is True
    assert ws.freeze_panes == "A2"
    assert isinstance(ws["A2"].value, (int, float))  # depth numeric
    assert isinstance(ws["B2"].value, (int, float))  # qc numeric
    header = [c.value for c in ws[1]]
    sbt_col = header.index("Soil Behaviour Type") + 1
    assert isinstance(ws.cell(row=2, column=sbt_col).value, str)  # text stays text

    raw = build_workbook(payload, now=_NOW)
    assert b"SHOULD NOT APPEAR" not in raw


def test_generic_builder_handles_multiple_tables():
    # A synthetic 2-table result -> 2 data sheets + Summary (proves genericity).
    payload = {
        "source_file": "x.CPT",
        "tables": [
            {"name": "Table A", "columns": [{"header": "n", "format": "0"}], "rows": [[1], [2]]},
            {"name": "Table B", "columns": [{"header": "v", "format": "0.0"}], "rows": [[3.5]]},
        ],
        "summary": {"Source file": "x.CPT"},
    }
    wb = load_workbook(io.BytesIO(build_workbook(payload, now=_NOW)))
    assert wb.sheetnames == ["Table A", "Table B", "Summary"]
    assert wb["Table B"]["A2"].value == 3.5


# --- CPT parity vs the pre-refactor golden output --------------------------
@pytest.mark.parametrize(
    "name,params",
    [
        ("default", {}),
        ("params", {"groundwater_level": 2.0, "soil_unit_weight": 18.0}),
    ],
)
def test_cpt_export_matches_golden(name, params):
    golden = _dump((_DATA_DIR / f"golden_cpt_{name}.xlsx").read_bytes())
    current = _dump(build_workbook(_result_object(params), now=_NOW))
    assert current == golden, f"CPT export drifted from golden ({name})"
