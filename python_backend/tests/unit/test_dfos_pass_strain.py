"""DFOS pass-strain calculator: determinism, contract shape, downsampling,
parameters, export tables. Synthetic arrays only (fast)."""

from __future__ import annotations

import numpy as np
import pytest

from app.core import config
from app.workspace.calculators.base import DatasetInput
from app.workspace.calculators.dataset_charts import downsample_xy
from app.workspace.calculators.dfos_pass_strain import DFOS_CALCULATOR, dfos_compute
from app.workspace.export.xlsx import build_workbook


def _dataset(
    n_t: int = 60, n_g: int = 400, speed: float = 2.0, seed: int = 1,
    artifact: bool = False, dead: tuple = (),
) -> DatasetInput:
    """Measurement rows are TARED (real-file convention); ``artifact`` puts the
    termination spike in the last three gages; ``dead`` gages are all-NaN."""
    rng = np.random.default_rng(seed)
    x = 0.08 + 0.0026 * np.arange(n_g) * 20  # ~0.05 m pitch -> ~20 m span
    tare = 900.0 * np.sin(x / 3.0)  # large locked-in baseline, like the real file
    dt = 0.12
    t = np.arange(n_t) * dt
    strain = np.empty((n_t, n_g), dtype=np.float32)
    for k in range(n_t):
        xl = -2.0 + speed * t[k]
        strain[k] = 100.0 * np.exp(-((x - xl) / 0.6) ** 2) + rng.normal(0, 0.5, n_g)
        if artifact:
            strain[k, -3:] = [11267.6, 7772.4, -15405.8]
    for g in dead:
        strain[:, g] = np.nan
    ts = (np.datetime64("2026-04-09T16:23:46.000") + (t * 1000).astype("timedelta64[ms]")).astype("datetime64[ms]")
    return DatasetInput(
        id="ds1", filename="pass_001.tsv", dataset_kind="strain_distributed",
        metadata={"sample_rate_hz": 1 / dt, "gage_pitch_mm": 52.0, "tare_name": "0409", "units": "microstrain"},
        arrays={"x_axis": x, "tare": tare, "strain": strain, "timestamps": ts,
                "timestamp_text": np.array([str(v).replace("T", " ") for v in ts])},
    )


def test_registration_binds_to_dataset_kind_not_extension():
    assert DFOS_CALCULATOR.required_dataset_kind == "strain_distributed"
    assert DFOS_CALCULATOR.required_extension == ""
    assert "run dfos pass strain" in DFOS_CALCULATOR.trigger_phrases
    ref = DFOS_CALCULATOR.reference.lower()
    assert "direct" in ref and "tare" in ref
    # Two SEPARATE exclusions with their reasons, both provisional; speed gating stated.
    assert f"lead-in {config.DFOS_LEADIN_EXCLUDE_M:g} m" in ref and "unbonded lead-in" in ref
    assert f"tail {config.DFOS_TAIL_EXCLUDE_M:g} m" in ref and "fibre termination artifact" in ref
    assert "provisional" in ref
    assert f"r2 >= {config.DFOS_SPEED_MIN_R2:g}" in ref and "not determinable" in ref


@pytest.fixture(autouse=True)
def default_trim(monkeypatch):
    monkeypatch.setattr(config, "DFOS_LEADIN_EXCLUDE_M", 1.10)
    monkeypatch.setattr(config, "DFOS_TAIL_EXCLUDE_M", 0.50)
    monkeypatch.setattr(config, "DFOS_SPEED_MIN_R2", 0.70)
    monkeypatch.setattr(config, "DFOS_BAND_WIDTH_M", 0.5)
    monkeypatch.setattr(config, "DFOS_BAND_HIGH_STRAIN_MICROSTRAIN", 2000.0)
    monkeypatch.setattr(config, "DFOS_SUBTRACT_TARE", False)


def test_recorded_strain_is_used_as_tare_relative_and_tare_not_subtracted_again():
    ds = _dataset()
    r = dfos_compute(ds, "pass.tsv", {})
    # Peak ~100 microstrain (the tared load), NOT 100 + tare (~900) nor 100 - tare.
    assert 95 < r.metadata["global_peak_microstrain"] < 106
    assert r.metadata["tare_subtracted"] is False
    assert "not subtracted again" in r.summary["Tare handling"]


def test_subtract_tare_flag_is_honoured(monkeypatch):
    monkeypatch.setattr(config, "DFOS_SUBTRACT_TARE", True)
    ds = _dataset()
    r = dfos_compute(ds, "pass.tsv", {})
    assert r.metadata["tare_subtracted"] is True
    # tare is ~900*sin(x/3): subtracting it shifts everything by hundreds.
    assert abs(r.metadata["global_peak_microstrain"] - 100) > 50


def test_two_exclusions_remove_lead_in_and_termination_but_export_keeps_full_fibre():
    ds = _dataset(n_g=400, artifact=True, dead=(151, 210, 211))
    x = ds.arrays["x_axis"]
    r = dfos_compute(ds, "pass_001.tsv", {})
    m = r.metadata
    lo = int(np.searchsorted(x, 1.10)); hi = int(np.searchsorted(x, x[-1] - 0.50, side="right"))
    assert m["leadin_exclude_m"] == 1.10 and m["leadin_exclude_reason"] == "unbonded lead-in"
    assert m["tail_exclude_m"] == 0.50 and m["tail_exclude_reason"] == "fibre termination artifact"
    assert m["leadin_excluded_gages"] == lo and m["tail_excluded_gages"] == 400 - hi
    assert m["analysed_gage_start"] == lo and m["analysed_gage_end_exclusive"] == hi
    assert m["analysed_x_min_m"] >= 1.10 and m["analysed_x_max_m"] <= x[-1] - 0.50
    # The artifact (11268 microstrain in the last gages) is NOT the global peak any more.
    assert m["global_peak_microstrain"] < 200 and lo <= m["global_peak_gage_index"] < hi
    assert m["global_peak_on_trim_boundary"] is False
    assert m["envelope_max_full_fibre_microstrain"] > 10000  # still visible in the full-fibre figure
    # Quality surfaced in the deterministic block, both exclusions stated separately with reasons.
    assert m["dead_gage_count"] == 3
    keys = list(r.summary)
    assert any(k.startswith("Lead-in excluded at head (m) - unbonded lead-in") for k in keys)
    assert any(k.startswith("Tail excluded at far end (m) - fibre termination artifact") for k in keys)
    assert r.summary["Lead-in excluded gages"] == lo and r.summary["Tail excluded gages"] == 400 - hi
    assert r.summary["Exclusion validation status"].startswith("PROVISIONAL")
    assert "NaN fraction (%)" in r.summary and "Analysed span (m)" in r.summary
    assert r.summary["Dead gages (all-NaN), whole fibre"] == 3
    # Notices: provisional exclusions first + peak-tracking info + quality.
    assert r.notices[0]["level"] == "provisional"
    assert "lead-in 1.1 m" in r.notices[0]["text"] and "tail 0.5 m" in r.notices[0]["text"]
    assert any("direct per-timestep maximum" in n["text"] for n in r.notices)
    # Envelope export keeps ALL gages, untrimmed, with region flags and the artifact values.
    env = r.tables[1]
    assert len(env["rows"]) == 400
    assert env["rows"][0][2].startswith("no (unbonded lead-in") and env["rows"][lo][2] == "yes"
    assert env["rows"][399][2].startswith("no (fibre termination artifact")
    assert env["rows"][397][4] == pytest.approx(11267.6, rel=1e-6)  # untrimmed precision in the export
    assert env["rows"][151][7] == "yes"  # dead gage flagged
    # Charts: analysed span only, excluded regions shaded, axis spans the full fibre.
    envc = r.charts[0]
    assert envc["x_range"] == [pytest.approx(0.08), pytest.approx(float(x[-1]))]
    assert [b["label"] for b in envc["bands"]] == ["lead-in 1.1 m excluded", "tail 0.5 m excluded"]
    xs = [v for v in envc["series"][0]["x"] if v is not None]
    assert min(xs) >= m["analysed_x_min_m"] - 1e-9 and max(xs) <= m["analysed_x_max_m"] + 1e-9


def test_band_profile_covers_full_fibre_and_is_exported():
    ds = _dataset(n_g=400, artifact=True, dead=(151,))
    x = ds.arrays["x_axis"]
    r = dfos_compute(ds, "pass.tsv", {})
    prof = r.metadata["band_profile"]
    assert len(prof) == int(np.ceil((x[-1] - x[0]) / 0.5))
    assert prof[0]["gage_from"] == 0 and prof[-1]["gage_to"] == 399  # full fibre incl. trimmed ends
    assert prof[0]["region"].startswith("lead-in (excluded") and prof[-1]["region"].startswith("tail (excluded")
    assert any(b["region"] == "analysed" for b in prof)
    assert prof[-1]["max_abs_strain"] == pytest.approx(15405.8, rel=1e-6)  # artifact shows in the tail band
    dead_band = next(b for b in prof if b["gage_from"] <= 151 <= b["gage_to"])
    assert dead_band["n_dead_gages"] == 1
    for b in prof:
        assert set(b) >= {"band", "x_from_m", "x_to_m", "gage_from", "gage_to", "n_gages", "n_dead_gages",
                          "max_abs_strain", "median_gage_max_abs_strain", "fraction_gages_above_level", "region"}
    # Chart values == profile values; sheet present in the export.
    bc = next(c for c in r.charts if c["id"] == "band_profile")
    assert bc["series"][0]["kind"] == "bar" and len(bc["series"][0]["x"]) == len(prof)
    assert bc["series"][0]["x"] == [b["x_from_m"] for b in prof]
    assert r.tables[3]["name"] == "Band profile" and len(r.tables[3]["rows"]) == len(prof)
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(build_workbook({"tables": r.tables, "summary": r.summary})))
    assert "Band profile" in wb.sheetnames
    assert wb["Band profile"].max_row == len(prof) + 1


def test_exclusions_from_config_and_inline_override(monkeypatch):
    ds = _dataset(n_g=400)
    x = ds.arrays["x_axis"]
    monkeypatch.setattr(config, "DFOS_LEADIN_EXCLUDE_M", 2.0)
    monkeypatch.setattr(config, "DFOS_TAIL_EXCLUDE_M", 1.0)
    r = dfos_compute(ds, "pass.tsv", {})
    assert r.metadata["leadin_exclude_m"] == 2.0 and r.metadata["tail_exclude_m"] == 1.0
    assert r.metadata["analysed_gage_start"] == int(np.searchsorted(x, 2.0))
    r2 = dfos_compute(ds, "pass.tsv", {"leadin_exclude_m": 0.3, "tail_exclude_m": 0.2})
    assert r2.metadata["leadin_exclude_m"] == 0.3 and r2.metadata["tail_exclude_m"] == 0.2
    assert r2.metadata["analysed_gage_count"] > r.metadata["analysed_gage_count"]
    r0 = dfos_compute(ds, "pass.tsv", {"leadin_exclude_m": 0.0, "tail_exclude_m": 0.0})
    assert r0.metadata["analysed_gage_count"] == 400 and r0.charts[0]["bands"] == []
    with pytest.raises(ValueError):
        dfos_compute(ds, "pass.tsv", {"leadin_exclude_m": 15.0, "tail_exclude_m": 15.0})


def test_speed_is_suppressed_everywhere_when_fit_is_not_credible():
    # A stationary "peak" (no moving load): the fit cannot be credible.
    ds = _dataset(n_t=40, n_g=300)
    ds.arrays["strain"][:] = 0.0
    ds.arrays["strain"][:, 150] = 100.0 + np.arange(40) * 0.01
    ds.arrays["strain"][:, 100] = 60.0
    r = dfos_compute(ds, "flat.tsv", {})
    m = r.metadata
    assert m["speed_credible"] is False
    assert m["implied_speed_m_s"] is None and m["implied_speed_kmh"] is None and m["direction"] is None
    assert m["speed_fit_r2"] is not None and "not determinable" in m["speed_status"]
    assert r.summary["Implied speed (m/s)"].startswith("not determinable (R²")
    assert r.summary["Implied speed (km/h)"] == "not determinable"
    assert r.summary["Direction of travel"] == "not determinable"
    assert "did not resolve a moving load" in r.summary["Load tracking"]
    assert "not determinable" in r.summary_text and "m/s" not in r.summary_text.split("not determinable")[0].split("implied speed")[-1]
    assert any(n["level"] == "warning" and "did not resolve a moving load" in n["text"] for n in r.notices)
    assert r.raw["implied_speed_m_s"] == "not determinable" and r.raw["direction"] == "not determinable"
    # No numeric speed anywhere in the summary values.
    for k, v in r.summary.items():
        if "speed" in k.lower() and "R2" not in k and "fraction" not in k.lower():
            assert not isinstance(v, (int, float)), (k, v)


def test_speed_is_reported_when_fit_clears_threshold():
    r = dfos_compute(_dataset(), "pass.tsv", {})
    m = r.metadata
    assert m["speed_fit_r2"] > 0.7 and m["speed_credible"] is True
    assert m["implied_speed_m_s"] == pytest.approx(2.0, rel=0.1) and m["direction"].startswith("+x")
    assert isinstance(r.summary["Implied speed (m/s)"], float)
    assert r.summary["Direction of travel"].startswith("+x")


def test_charts_are_downsampled_to_at_most_2000_points():
    ds = _dataset(n_t=50, n_g=7795)
    r = dfos_compute(ds, "pass.tsv", {})
    ids = [c["id"] for c in r.charts]
    assert ids == ["envelope", "band_profile", "influence", "load_position"]
    for c in r.charts:
        for s in c["series"]:
            assert len(s["x"]) <= 2000 and len(s["y"]) <= 2000 and len(s["x"]) == len(s["y"])
    env = r.charts[0]["series"][0]
    assert 7000 < env["n_source"] < 7795 and len(env["x"]) == 2000  # analysed span (lead-in + tail excluded)
    # Peak survives the downsampling (extreme-preserving bins).
    assert max(v for v in env["y"] if v is not None) == pytest.approx(r.metadata["envelope_max_microstrain"], rel=1e-5)  # 6 sig. digits in chart payload
    # Full precision stays in the export tables.
    assert len(r.tables[1]["rows"]) == 7795


def test_downsample_keeps_extremes_and_passes_small_arrays_through():
    x = np.arange(10.0)
    y = np.zeros(10)
    xs, ys = downsample_xy(x, y, max_points=20)
    assert xs.shape == (10,) and ys.shape == (10,)
    x = np.arange(10000.0)
    y = np.zeros(10000)
    y[1234] = 50.0
    y[8765] = -80.0
    xs, ys = downsample_xy(x, y, max_points=100)
    assert xs.shape == (100,)
    assert 50.0 in ys and -80.0 in ys


def test_influence_gage_parameter_selects_nearest_gage():
    ds = _dataset()
    r = dfos_compute(ds, "pass.tsv", {"gage_x": 10.0})
    assert abs(r.metadata["influence_gage_x_m"] - 10.0) < 0.06
    assert r.charts[2]["title"].startswith("Influence line at gage x = ")
    r2 = dfos_compute(ds, "pass.tsv", {"peak_fraction": 0.5})
    assert r2.metadata["loaded_timesteps"] < r.metadata["loaded_timesteps"]


def test_export_workbook_has_full_precision_tables():
    import io
    from openpyxl import load_workbook

    ds = _dataset(n_t=30, n_g=300)
    r = dfos_compute(ds, "pass_001.tsv", {})
    payload = {"tables": r.tables, "summary": r.summary, "source_file": "pass_001.tsv"}
    wb = load_workbook(io.BytesIO(build_workbook(payload)))
    assert wb.sheetnames == ["Peak per timestep", "Envelope", "Influence line", "Band profile", "Summary"]
    env = wb["Envelope"]
    assert env.max_row == 301  # header + 300 gages
    val = env.cell(row=2, column=5).value  # "Max strain" column (col 3 is the span flag, col 4 the tare)
    assert isinstance(val, float) and val == pytest.approx(r.tables[1]["rows"][0][4], rel=1e-15)  # untruncated float
    summ = {row[0].value: row[1].value for row in wb["Summary"].iter_rows(min_row=2)}
    assert "Global peak strain (microstrain)" in summ
    assert summ["Exclusion validation status"].startswith("PROVISIONAL")
    assert summ["Lead-in excluded at head (m) - unbonded lead-in (provisional)"] == 1.1
    assert summ["Tail excluded at far end (m) - fibre termination artifact (provisional)"] == 0.5
    assert summ["Peak tracking"] == "direct maximum per timestep (not a fitted model)"


def test_all_nan_dataset_raises_value_error():
    ds = _dataset(n_t=5, n_g=10)
    ds.arrays["strain"] = np.full((5, 10), np.nan, dtype=np.float32)
    with pytest.raises(ValueError):
        dfos_compute(ds, "x.tsv", {})
