"""Thread-wide highlight export (HIGHLIGHTS_ENABLED) + thread-delete cascade.

Deterministic: fake collections, route coroutines called directly, workbook
bytes re-opened with openpyxl. Covers empty thread (both formats), ownership,
ordering by message then offset, context window (sentence/word trimming,
ellipses, markdown cleanup), newline-in-note hard breaks, markdown escaping,
xlsx headers/rows/wrap, flag-off route absence, and the cascade.
"""
import io
from datetime import datetime
from types import SimpleNamespace

import pytest
from bson import ObjectId
from fastapi import FastAPI, HTTPException
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core import config
from app.routers import highlights as hl
from app.routers import threads as threads_mod
from app.services import highlights_export as ex
from tests.unit.test_highlights import FakeCollection

pytestmark = pytest.mark.unit

OWNER = SimpleNamespace(id="user-owner")
OTHER = SimpleNamespace(id="user-other")
THREAD = "thread_abc"
T0 = datetime(2026, 8, 17, 12, 0, 0)

MSG1 = ObjectId()
MSG2 = ObjectId()
CONTENT1 = (
    "### Definition of Bearing Capacity\n\n"
    "Bearing capacity is the maximum load per unit area that a soil can support without shear failure. "
    "It is a cornerstone of foundation design.\n\n"
    "1.  **Failure Modes**: Bearing capacity is often limited by two distinct phenomena:\n"
    "*   **Shear failure** of the soil mass.\n"
    "*   **Excessive settlement** under load."
)
# Highlight #2 spans "Failure Modes**: Bearing capacity" in the source (bold
# delimiters inside the range, as the frontend anchoring produces).
S2 = CONTENT1.index("Failure Modes")
E2 = CONTENT1.index("Bearing capacity is often") + len("Bearing capacity")
CONTENT2 = "Liquefaction occurs when saturated sand loses strength under cyclic loading. See `qc` values."


def _h(msg, start, end, text, colour="yellow", note="", created=T0, _id=None):
    return {"_id": _id or ObjectId(), "threadId": THREAD, "messageId": str(msg), "userId": OWNER.id,
            "startOffset": start, "endOffset": end, "selectedText": text, "colour": colour, "note": note,
            "createdAt": created, "updatedAt": created}


@pytest.fixture
def db(monkeypatch):
    conversations = FakeCollection([
        {"_id": ObjectId(), "userId": OWNER.id, "threadId": THREAD, "name": "Bearing capacity Q&A"},
        {"_id": ObjectId(), "userId": OTHER.id, "threadId": "thread_other", "name": "o"},
    ])
    messages = FakeCollection([
        {"_id": MSG1, "threadId": THREAD, "userId": OWNER.id, "role": "assistant", "content": CONTENT1, "createdAt": T0},
        {"_id": MSG2, "threadId": THREAD, "userId": OWNER.id, "role": "assistant", "content": CONTENT2, "createdAt": datetime(2026, 8, 17, 12, 5)},
    ])
    highlights = FakeCollection()
    monkeypatch.setattr(hl, "conversations_collection", conversations)
    monkeypatch.setattr(hl, "messages_collection", messages)
    monkeypatch.setattr(hl, "highlights_collection", highlights)
    monkeypatch.setattr(config, "HIGHLIGHTS_ENABLED", True)
    return SimpleNamespace(conversations=conversations, messages=messages, highlights=highlights)


# --- context window ----------------------------------------------------------

def test_context_trims_to_sentence_boundaries_and_ellipsises():
    s = CONTENT1.index("Bearing capacity is often")
    e = s + len("Bearing capacity is often")
    before, after = ex.context_window(CONTENT1, s, e, radius=80)
    # before: starts at a sentence/paragraph boundary inside the window, markdown stripped
    assert before == "…Failure Modes:"
    # after: cut at the end of the current sentence-ish run, list markers stripped
    assert after == "limited by two distinct phenomena: Shear failure of the soil mass.…"


def test_context_no_ellipsis_at_message_edges_and_word_boundary_fallback():
    # span at the very start: nothing omitted before -> no leading ellipsis;
    # after is cut at a word boundary (no sentence end in a 30-unit window)
    assert ex.context_window(CONTENT2, 0, len("Liquefaction"), radius=30) == ("", "occurs when saturated sand…")
    # tiny window, no sentence end: word boundaries on the far edges, ellipsised
    assert CONTENT2[40:47] == "loses s"
    assert ex.context_window(CONTENT2, 40, 47, radius=12) == ("…sand", "trength…")
    # window reaching the end of the message: no trailing ellipsis; code
    # backticks stripped from the context
    assert ex.context_window(CONTENT2, len(CONTENT2) - 10, len(CONTENT2) - 5, radius=50) == ("…See q", "lues.")


def test_context_offsets_are_utf16_units():
    content = "\U0001F600 friction angle of sand"
    start = content.index("friction") + 1  # JS index (astral char = 2 units)
    before, after = ex.context_window(content, start, start + 8)
    assert before == "\U0001F600"
    assert after == "angle of sand"


def test_join_context_spacing():
    assert ex._join_context("…before", "**s**", ", after…") == "…before **s**, after…"
    assert ex._join_context("open (", "**s**", ") close") == "open (**s**) close"
    assert ex._join_context("", "**s**", "after") == "**s** after"
    assert ex._join_context("before", "**s**", "") == "before **s**"


def test_clean_context_strips_markdown_and_collapses_whitespace():
    raw = "## Heading\n\n> quoted **bold** and _em_ and `code` [link](http://x) | cell |\n- item\n2. two  \n\n---\n"
    assert ex._clean_context(raw) == "Heading quoted bold and em and code link cell item two"


# --- markdown ----------------------------------------------------------------

def test_escape_markdown():
    assert ex.escape_markdown("a*b_c`d[e]f<g>h#i|j~k\\l") == "a\\*b\\_c\\`d\\[e\\]f\\<g\\>h\\#i\\|j\\~k\\\\l"
    assert ex.escape_markdown("- item\n+ plus\n1. one\n2) two") == "\\- item\n\\+ plus\n1\\. one\n2\\) two"
    assert ex.escape_markdown("plain text 3.5 kPa") == "plain text 3.5 kPa"


def test_markdown_orders_by_message_then_offset_with_context_and_notes():
    rows = ex.build_rows(
        [
            _h(MSG2, 0, 12, "Liquefaction", colour="blue", created=datetime(2026, 8, 17, 12, 1)),
            _h(MSG1, S2, E2, "Failure Modes: Bearing capacity", colour="green", note="exam Q3\nsee *Terzaghi*", created=datetime(2026, 8, 17, 12, 2)),
            _h(MSG1, 36, 72, "Bearing capacity is the maximum load", created=datetime(2026, 8, 17, 12, 3)),
        ],
        [{"_id": MSG1, "content": CONTENT1}, {"_id": MSG2, "content": CONTENT2}],
    )
    assert [r.selected_text for r in rows] == [
        "Bearing capacity is the maximum load",      # message 1, offset 36
        "Failure Modes: Bearing capacity",           # message 1, later offset
        "Liquefaction",                              # message 2
    ]
    md = ex.build_markdown("Bearing capacity Q&A", rows, now=T0)
    assert md.startswith("# Highlights — Bearing capacity Q&A\n\n_3 highlights · exported 2026-08-17 12:00_\n")
    assert md.index("## 1.") < md.index("## 2.") < md.index("## 3.")
    # blockquote + note with hard breaks (two trailing spaces) + escaped emphasis chars
    assert "> Failure Modes: Bearing capacity\n\n**Note:** exam Q3  \nsee \\*Terzaghi\\*\n" in md
    # context line: prose before, span in bold, prose after
    assert "_Context:_ …It is a cornerstone of foundation design. **Failure Modes: Bearing capacity** is often limited by two distinct phenomena:" in md
    assert "_Colour: green · 2026-08-17 12:02_" in md
    # a highlight without a note has no Note line
    sec1 = md[md.index("## 1."):md.index("## 2.")]
    assert "**Note:**" not in sec1


def test_markdown_escapes_highlighted_text_and_multiline_blockquote():
    rows = ex.build_rows([_h(MSG1, 0, 5, "# not a heading\n\n* not a bullet [x](y)")], [{"_id": MSG1, "content": CONTENT1}])
    md = ex.build_markdown("t", rows, now=T0)
    assert "> \\# not a heading\n>\n> \\* not a bullet \\[x\\](y)\n" in md


def test_markdown_empty_thread_is_valid():
    md = ex.build_markdown("Empty *thread*", [], now=T0)
    assert md == "# Highlights — Empty \\*thread\\*\n\n_0 highlights · exported 2026-08-17 12:00_\n\n_No highlights in this thread._\n"


def test_markdown_missing_message_sorts_last_without_context():
    rows = ex.build_rows([_h(ObjectId(), 0, 3, "abc"), _h(MSG1, 0, 3, "###")], [{"_id": MSG1, "content": CONTENT1}])
    assert [r.selected_text for r in rows] == ["###", "abc"]
    assert rows[1].context_available is False
    assert "_Context:_ (message no longer available)" in ex.build_markdown("t", rows, now=T0)


# --- xlsx --------------------------------------------------------------------

def _open(xlsx_bytes):
    return load_workbook(io.BytesIO(xlsx_bytes))


def test_xlsx_headers_rows_context_and_wrap():
    rows = ex.build_rows(
        [_h(MSG1, 36, 72, "Bearing capacity is the maximum load", note="line one\nline two", colour="pink")],
        [{"_id": MSG1, "content": CONTENT1}],
    )
    wb = _open(ex.build_xlsx("Bearing capacity Q&A", rows, now=T0))
    ws = wb["Highlights"]
    assert [c.value for c in ws[1]] == ["Date", "Colour", "Highlighted text", "Note", "Context", "Thread title"]
    row = [c.value for c in ws[2]]
    assert row[0] == "2026-08-17 12:00" and row[1] == "pink"
    assert row[2] == "Bearing capacity is the maximum load"
    assert row[3] == "line one\nline two"          # newline preserved in the cell
    assert row[4].startswith("Definition of Bearing Capacity [Bearing capacity is the maximum load] per unit area")
    assert row[5] == "Bearing capacity Q&A"
    assert ws.max_row == 2
    # wrap-text post-pass on text/note/context columns
    for col in ("C", "D", "E"):
        assert ws[f"{col}2"].alignment.wrap_text is True
        assert ws.column_dimensions[col].width == ex.XLSX_WRAP_WIDTH
    assert ws["A2"].alignment.wrap_text is not True
    # generic summary sheet from the GeoPilot builder is present and untouched in shape
    summ = wb["Summary"]
    assert [c.value for c in summ[1]] == ["Field", "Value"]
    assert ("Thread", "Bearing capacity Q&A") in [(r[0].value, r[1].value) for r in summ.iter_rows(min_row=2)]


def test_xlsx_empty_thread_is_valid():
    wb = _open(ex.build_xlsx("t", [], now=T0))
    ws = wb["Highlights"]
    assert [c.value for c in ws[1]] == ex.XLSX_COLUMNS
    assert ws.max_row == 1


# --- route -------------------------------------------------------------------

async def test_export_route_md_and_xlsx(db):
    await db.highlights.insert_one(_h(MSG2, 0, 12, "Liquefaction", colour="blue"))
    await db.highlights.insert_one(_h(MSG1, 36, 72, "Bearing capacity is the maximum load", note="n1\nn2"))
    md = await hl.export_highlights(THREAD, format="md", current_user=OWNER)
    assert md.status_code == 200
    assert md.media_type == "text/markdown; charset=utf-8"
    assert md.headers["content-disposition"].startswith('attachment; filename="highlights_Bearing_capacity_Q_A_')
    text = md.body.decode("utf-8")
    assert text.index("Bearing capacity is the maximum load") < text.index("> Liquefaction")
    assert "**Note:** n1  \nn2" in text
    x = await hl.export_highlights(THREAD, format="XLSX", current_user=OWNER)  # case-insensitive
    assert x.media_type == ex.XLSX_MEDIA_TYPE and x.body[:2] == b"PK"
    ws = _open(x.body)["Highlights"]
    assert [c.value for c in ws[1]] == ex.XLSX_COLUMNS and ws.max_row == 3
    assert ws["C2"].value == "Bearing capacity is the maximum load" and ws["C3"].value == "Liquefaction"


async def test_export_route_empty_thread_and_bad_format(db):
    md = await hl.export_highlights(THREAD, format="md", current_user=OWNER)
    assert md.status_code == 200 and b"_No highlights in this thread._" in md.body
    x = await hl.export_highlights(THREAD, format="xlsx", current_user=OWNER)
    assert _open(x.body)["Highlights"].max_row == 1
    with pytest.raises(HTTPException) as e:
        await hl.export_highlights(THREAD, format="pdf", current_user=OWNER)
    assert e.value.status_code == 422


async def test_export_route_ownership(db):
    await db.highlights.insert_one(_h(MSG1, 36, 72, "Bearing capacity is the maximum load"))
    with pytest.raises(HTTPException) as e:
        await hl.export_highlights(THREAD, format="md", current_user=OTHER)
    assert e.value.status_code == 404 and e.value.detail == "Thread not found"
    with pytest.raises(HTTPException) as e:
        await hl.export_highlights("null", format="md", current_user=OWNER)
    assert e.value.status_code == 400


# --- flag off ----------------------------------------------------------------

def test_export_route_absent_when_flag_off_present_when_on(monkeypatch):
    path = "/api/assistants/threads/{thread_id}/highlights/export"
    monkeypatch.setattr(config, "HIGHLIGHTS_ENABLED", False)
    app = FastAPI(); hl.register(app)
    assert path not in app.openapi()["paths"]
    assert TestClient(app).get("/api/assistants/threads/t1/highlights/export?format=md").status_code == 404
    monkeypatch.setattr(config, "HIGHLIGHTS_ENABLED", True)
    app = FastAPI(); hl.register(app)
    assert sorted(app.openapi()["paths"][path]) == ["get"]
    assert TestClient(app).get("/api/assistants/threads/t1/highlights/export?format=md").status_code == 401


# --- thread-delete cascade ---------------------------------------------------

async def test_thread_delete_cascades_to_highlights_only_for_that_thread_and_user(monkeypatch):
    conversations = FakeCollection([
        {"_id": ObjectId(), "userId": OWNER.id, "threadId": THREAD, "name": "t"},
        {"_id": ObjectId(), "userId": OWNER.id, "threadId": "thread_keep", "name": "k"},
        {"_id": ObjectId(), "userId": OTHER.id, "threadId": THREAD, "name": "same id, other user"},
    ])
    messages = FakeCollection([{"_id": MSG1, "threadId": THREAD, "userId": OWNER.id, "role": "assistant", "content": "x"}])
    files = FakeCollection()
    highlights = FakeCollection([
        _h(MSG1, 0, 1, "x"),                                    # this thread, this user -> gone
        _h(MSG1, 0, 1, "x"),                                    # this thread, this user -> gone
        {**_h(MSG1, 0, 1, "x"), "threadId": "thread_keep"},     # other thread, same user -> stays
        {**_h(MSG1, 0, 1, "x"), "userId": OTHER.id},            # same thread id, OTHER user -> stays
    ])
    # the route's delete_many/delete_one/count_documents on the fakes
    async def delete_many(self, flt):
        gone = [k for k, d in self.docs.items() if all(d.get(a) == b for a, b in flt.items())]
        for k in gone: del self.docs[k]
        return SimpleNamespace(deleted_count=len(gone))
    async def count_documents(self, flt):
        return sum(1 for d in self.docs.values() if all(d.get(a) == b for a, b in flt.items() if not isinstance(b, dict)))
    for c in (conversations, messages, files, highlights):
        monkeypatch.setattr(type(c), "delete_many", delete_many, raising=False)
        monkeypatch.setattr(type(c), "count_documents", count_documents, raising=False)
    monkeypatch.setattr(threads_mod, "conversations_collection", conversations)
    monkeypatch.setattr(threads_mod, "messages_collection", messages)
    monkeypatch.setattr(threads_mod, "files_collection", files)
    monkeypatch.setattr(threads_mod, "highlights_collection", highlights)

    out = await threads_mod.delete_thread(SimpleNamespace(threadId=THREAD), current_user=OWNER)
    # response shape unchanged: no new key
    assert out == {"success": True, "message": "Thread deleted successfully", "deleted_messages": 1, "deleted_thread_documents": 0}
    # zero orphan highlight docs for the deleted thread+user; the other two survive
    assert await highlights.count_documents({"userId": OWNER.id, "threadId": THREAD}) == 0
    remaining = list(highlights.docs.values())
    assert sorted((d["userId"], d["threadId"]) for d in remaining) == [(OTHER.id, THREAD), (OWNER.id, "thread_keep")]
