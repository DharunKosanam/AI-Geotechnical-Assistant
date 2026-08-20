"""Traffic-load monitoring calculator + event-detection strategy: determinism,
provisional status everywhere, direction from channel order, swappable
strategy, plausibility warning. Small synthetic arrays only."""

from __future__ import annotations

import numpy as np
import pytest

from app.core import config
from app.workspace.calculators.base import DatasetInput
from app.workspace.calculators.event_detection import (
    STRATEGIES,
    DetectionParams,
    block_percentile_baseline,
    detect_percentile_mad,
    get_strategy,
)
from app.workspace.calculators.traffic_load_monitoring import (
    PROVISIONAL_TEXT,
    TRAFFIC_CALCULATOR,
    traffic_compute,
)


def _dataset(n_events: int = 12, hours: float = 2.0, forward: bool = True, seed: int = 3) -> DatasetInput:
    rng = np.random.default_rng(seed)
    n = int(hours * 3600 * 10)
    t = np.arange(n) * 0.1
    names = ["A_kPa", "B_kPa", "C_kPa", "D_kPa"]
    base = np.array([11.6, 9.9, 10.2, 7.5])
    p = base[None, :] + rng.normal(0, 0.1, (n, 4))
    centres = np.linspace(300, hours * 3600 - 300, n_events)
    for k, tc in enumerate(centres):
        amp = 10 + 3 * k
        for c in range(4):
            lag = (c if forward else 3 - c) * 0.4
            p[:, c] += amp * np.exp(-((t - tc - lag) / 0.5) ** 2)
    ts = (np.datetime64("2024-05-10T00:00:00.000") + (t * 1000).astype("timedelta64[ms]")).astype("datetime64[ms]")
    return DatasetInput(
        id="ds", filename="2024-05-10.dat", dataset_kind="pressure_timeseries",
        metadata={"channel_names": names, "sample_rate_hz": 10.0, "first_timestamp": str(ts[0]), "last_timestamp": str(ts[-1])},
        arrays={"timestamps": ts, "record": np.arange(n), "pressure": p.astype(np.float32)},
    )


def test_registration_binds_to_dataset_kind_and_states_provisional():
    assert TRAFFIC_CALCULATOR.required_dataset_kind == "pressure_timeseries"
    assert TRAFFIC_CALCULATOR.required_extension == ""
    assert "run traffic load monitoring" in TRAFFIC_CALCULATOR.trigger_phrases
    ref = TRAFFIC_CALCULATOR.reference
    assert "PROVISIONAL" in ref and "pending validation" in ref
    assert "percentile_mad" in ref and "6" in ref  # exact method + parameters named


def test_detects_events_deterministically_with_direction():
    ds = _dataset(n_events=12, forward=True)
    r1 = traffic_compute(ds, "2024-05-10.dat", {})
    r2 = traffic_compute(ds, "2024-05-10.dat", {})
    assert r1.metadata == r2.metadata and r1.tables == r2.tables and r1.segments == r2.segments and r1.charts == r2.charts
    assert r1.metadata["n_events"] == 12
    assert len(r1.segments) == 12
    assert all(s["direction"] == "A_kPa → D_kPa" for s in r1.segments)
    back = traffic_compute(_dataset(n_events=5, forward=False), "x.dat", {})
    assert all(s["direction"] == "D_kPa → A_kPa" for s in back.segments)
    # Largest event is the last (amplitude grows with k).
    assert r1.metadata["largest_events"][0]["index"] == 12
    seg = r1.segments[0]
    assert set(seg) >= {"index", "label", "start", "end", "duration_s", "peak_sum_kpa", "peak_kpa", "channel_order", "direction"}
    assert seg["label"].startswith("Event 1 · ")


def test_provisional_status_is_visible_in_notices_summary_tables_and_reference():
    r = traffic_compute(_dataset(), "2024-05-10.dat", {})
    assert r.notices[0]["level"] == "provisional"
    assert PROVISIONAL_TEXT in r.notices[0]["text"]
    assert "percentile_mad" in r.notices[0]["text"] and "6" in r.notices[0]["text"]
    assert r.summary["Validation status"].startswith("PROVISIONAL")
    assert "PROVISIONAL" in r.summary["Method / Standard reference"]
    assert r.tables[0]["name"] == "Events (provisional)"
    assert r.tables[0]["rows"][0][-1] == "PROVISIONAL - not validated"
    assert "not validated" in r.summary_text
    assert r.metadata["validation_status"].startswith("PROVISIONAL")
    # Nothing in the payload calls the counts validated.
    import json
    blob = json.dumps({"s": r.summary, "m": r.metadata, "n": r.notices, "t": r.summary_text}, default=str).lower()
    assert "not validated" in blob and " validated counts" not in blob


def test_parameters_from_config_and_inline_overrides(monkeypatch):
    monkeypatch.setattr(config, "INSTRUMENT_EVENT_MAD_MULTIPLIER", 6.0)
    ds = _dataset(n_events=6)
    base = traffic_compute(ds, "x.dat", {})
    assert base.metadata["detection"]["mad_multiplier"] == 6.0
    hi = traffic_compute(ds, "x.dat", {"mad_multiplier": 200.0})
    assert hi.metadata["detection"]["mad_multiplier"] == 200.0
    assert hi.metadata["n_events"] < base.metadata["n_events"]
    monkeypatch.setattr(config, "INSTRUMENT_EVENT_MAD_MULTIPLIER", 4.0)
    cfg = traffic_compute(ds, "x.dat", {})
    assert cfg.metadata["detection"]["mad_multiplier"] == 4.0
    assert "MAD multiplier" in cfg.summary and cfg.summary["MAD multiplier"] == 4.0


def test_implausible_count_yields_warning_notice(monkeypatch):
    ds = _dataset(n_events=6)
    r = traffic_compute(ds, "x.dat", {"mad_multiplier": 1e6})  # nothing crosses
    assert r.metadata["n_events"] == 0 and r.metadata["plausible"] is False
    assert any(n["level"] == "warning" and "Implausible" in n["text"] for n in r.notices)


def test_strategy_registry_is_swappable():
    assert "percentile_mad" in STRATEGIES and get_strategy("percentile_mad") is detect_percentile_mad
    with pytest.raises(ValueError):
        get_strategy("nope")
    ds = _dataset(n_events=4)
    called = {}

    def fake(pressure, t_s, params):
        called["n"] = pressure.shape[0]
        n, c = pressure.shape
        return {"events": [], "baseline": np.zeros((n, c)), "residual": pressure.astype(float),
                "noise": np.ones(c), "threshold": np.ones(c)}

    STRATEGIES["fake"] = fake
    try:
        r = traffic_compute(ds, "x.dat", {"strategy": "fake"})
        assert called["n"] == ds.arrays["pressure"].shape[0]
        assert r.metadata["detection"]["strategy"] == "fake" and r.metadata["n_events"] == 0
    finally:
        del STRATEGIES["fake"]
    with pytest.raises(ValueError):
        traffic_compute(ds, "x.dat", {"strategy": "missing"})


def test_block_percentile_baseline_tracks_slow_drift():
    t = np.arange(0, 3600, 0.1)
    y = 10 + 0.001 * t + np.random.default_rng(0).normal(0, 0.05, t.size)
    b = block_percentile_baseline(y, t, 20.0, 300.0)
    assert b.shape == y.shape
    # Interior (block centres are interpolated; the ends are clamped).
    i0, i1 = 6000, 30000  # t = 600 s and 3000 s
    assert abs((b[i1] - b[i0]) - 0.001 * (t[i1] - t[i0])) < 0.15
    # Short series: constant percentile.
    bs = block_percentile_baseline(y[:100], t[:100], 20.0, 300.0)
    assert np.allclose(bs, bs[0])


def test_charts_are_small_and_export_has_provisional_rows():
    import io
    from openpyxl import load_workbook
    from app.workspace.export.xlsx import build_workbook

    r = traffic_compute(_dataset(n_events=8, hours=3), "2024-05-10.dat", {})
    for c in r.charts:
        for s in c["series"]:
            assert len(s["x"]) <= 2000
    assert r.charts[0]["series"][0]["kind"] == "bar" and len(r.charts[0]["series"][0]["x"]) == 24
    wb = load_workbook(io.BytesIO(build_workbook({"tables": r.tables, "summary": r.summary})))
    assert wb.sheetnames == ["Events (provisional)", "Hourly histogram", "Peak distribution", "Channels", "Summary"]
    summ = {row[0].value: row[1].value for row in wb["Summary"].iter_rows(min_row=2)}
    assert summ["Validation status"].startswith("PROVISIONAL")
